from django.shortcuts import render, redirect
from rest_framework.views import APIView
from django.http import JsonResponse, HttpResponse
from rest_framework.response import Response
import pandas as pd
from io import BytesIO
from tempfile import NamedTemporaryFile
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# 使用pandas處理
def export(file_data_form):
    data = file_data_form.get('Excel')
    
    # 讀取 Excel 檔案，返回一個 DataFrame
    df = pd.read_excel(data, engine='openpyxl')
    
    if df.empty:
        titles = ['訂單編號', '客戶名稱', '訂購日期', '數量', '價格']
        new_data = [
            ['4', '客戶D', '2024/01/04', 15, 750],
            ['5', '客戶E', '2024/01/05', 8, 400],
            ['6', '客戶F', '2024/01/06', 30, 1500]
        ]
        
        df = pd.DataFrame(new_data, columns=titles)
    else:
        # 確保排序條件是字串
        df['訂單編號'] = df['訂單編號'].astype(str)
        titles = df.columns.tolist() # ['訂單編號', '客戶名稱', '訂購日期', '數量', '價格']

        new_data = [
            ['4', '客戶D', '2024/01/04', 15, 750],
            ['5', '客戶E', '2024/01/05', 8, 400],
            ['6', '客戶F', '2024/01/06', 30, 1500]
        ]

        # 將新的資料轉換為 DataFrame
        new_df = pd.DataFrame(new_data, columns=titles)
        #合併資料
        df = pd.concat([df, new_df], ignore_index=True)
        df = df.sort_values(by='訂單編號')
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='訂單資料')
        workbook = writer.book
        sheet = workbook['訂單資料']

        # 設置標題行格式 (加粗並置中)
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # 設置整個表格的對齊方式 (置中)
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=len(df.columns), max_row=len(df) + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.ms-excel')

    return response

# 使用openpyxl處理
def export_by_openpyxl(file_data_form):
    data = file_data_form.get('Excel')
    wb = openpyxl.load_workbook(data)
    sheet = wb.active
    if sheet.max_row <= 1:
        # 如果沒有資料，創建一個新的標題和資料
        titles = ['訂單編號', '客戶名稱', '訂購日期', '數量', '價格']
        new_data = [
            ['4', '客戶D', '2024/01/04', 15, 750],
            ['5', '客戶E', '2024/01/05', 8, 400],
            ['6', '客戶F', '2024/01/06', 30, 1500]
        ]
        
        # 創建標題行
        for col_num, title in enumerate(titles, 1):
            sheet.cell(row=1, column=col_num, value=title)
        
        # 新增資料行
        for row_num, row in enumerate(new_data, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
    else:
        # 讀取現有的標題
        titles = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

        # 新增資料
        new_data = [
            ['4', '客戶D', '2024/01/04', 15, 750],
            ['5', '客戶E', '2024/01/05', 8, 400],
            ['6', '客戶F', '2024/01/06', 30, 1500]
        ]
        
        # 找到最後一行，從那裡開始新增資料
        last_row = sheet.max_row + 1
        for row in new_data:
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=last_row, column=col_num, value=value)
            last_row += 1

        # 重新排序資料（以訂單編號排序）
        rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column))
        rows.sort(key=lambda x: str(x[0].value))  # 假設 '訂單編號' 在第一欄

        # 清空原本資料區域，並重寫排序後的資料
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col, value=None)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, cell in enumerate(row, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell.value)

    # 設置標題行格式 (加粗並置中)
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    # 設置整個表格的對齊方式 (置中)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 儲存並返回 Excel 檔案
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # response = HttpResponse(output, content_type='application/vnd.ms-excel')
    response = Response(output.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="updated_orders.xlsx"'

    return response

    
    # with NamedTemporaryFile(delete=False) as tmp:
    #     wb.save(tmp.name)
    #     output = BytesIO(tmp.read())
    # export_file = HttpResponse(output, content_type='application/vnd.ms-excel')

    # return export_file


class Excel01(APIView):
    template_name = 'excel01.html'

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            # 如果未登錄，重定向到登錄頁面
            return redirect('login')  
        data = {}
        return render(request, self.template_name, data)
    
    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        file_data_form = request.FILES
        export_file = export(file_data_form)

        return export_file