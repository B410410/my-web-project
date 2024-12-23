from django.shortcuts import render
from django.views.generic import TemplateView
from django.http import JsonResponse,HttpResponse
from erp.viewspublic import *
import calendar
from tempfile import NamedTemporaryFile
from io import BytesIO
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime,timedelta
from openpyxl.styles import PatternFill, Font
import re
import random
"""
Database Process
"""


class Statement(Tables):

    def __init__(self):
        super().__init__()

    def stmt_select_erp_asset(self, year):
        t1 = 'erp_asset'
        t2 = 'erp_asset_installment'
        t3 = 'erp_account'
        stmt = f'''
            SELECT 
                {t1}.asset_no, 
                {t1}.account_no, 
                {t1}.asset_vendor, 
                {t1}.asset_name, 
                {t1}.depart_no, 
                {t1}.credit_account_no,
                {t1}.asset_qty, 
                {t1}.asset_unit, 
                {t1}.asset_date, 
                {t1}.asset_amount, 
                {t1}.number_installment, 
                {t1}.asset_residual,
                {t1}.open_grand_total,
                COALESCE({t2}.installment_date, NULL) AS installment_date, 
                COALESCE({t2}.balance_amount, NULL) AS balance_amount, 
                COALESCE({t2}.installment_amount, NULL) AS installment_amount, 
                COALESCE({t2}.resi_balance_amount, NULL) AS resi_balance_amount, 
                COALESCE({t2}.resi_installment_amount, NULL) AS resi_installment_amount,
                {t3}.account_name,
                COALESCE(last_year.installment_amount_total, 0) AS initial_grand_total  -- 前一年的installment_amount總和(期初累折)
            FROM 
                {t1}
            LEFT JOIN 
                {t2} ON {t1}.asset_no = {t2}.asset_no AND EXTRACT(YEAR FROM {t2}.installment_date) = {year}
            LEFT JOIN
                (SELECT
                    asset_no,
                    SUM(installment_amount) AS installment_amount_total
                FROM
                    {t2}
                WHERE 
                    EXTRACT(YEAR FROM installment_date) < {year}
                GROUP BY 
                    asset_no) AS last_year
                ON {t1}.asset_no = last_year.asset_no
            JOIN 
                {t3} ON {t1}.account_no = {t3}.account_no 
            WHERE
                ({t1}.is_installment OR {t1}.is_resi_installment OR ({t1}.asset_residual > 0))
            ORDER BY
                {t2}.installment_date;
        '''
        return stmt




"""
Main Process
"""

def add_commas_to_integer(number):
    number_str = str(number)
    # 在整數部分添加千分位分隔符
    return re.sub(r'(\d)(?=(\d{3})+(?!\d))', r'\1,', number_str)


def acc3003_main(data_form):
    input_year  = data_form.get('fmText01')
    input_month  = data_form.get('fmText02')
    data_d = {}
    dbh = ErpDb()

    try:
        erp_asset_data = dbh.erpdb_fetchall(Statement().stmt_select_erp_asset(int(input_year)))
    except Exception as e:
        print('資料查詢錯誤:', e)
        return {}
    finally:
        dbh.close()

    # 整理資料
    for data in erp_asset_data:
        asset_no = data['asset_no'] 
        
        # 計算期初累折
        balance_amount = int(data.get('balance_amount', 0)) if data.get('balance_amount') is not None else 0
        resi_balance_amount = int(data['resi_balance_amount']) if data['resi_balance_amount'] is not None else 0
        initial_grand_total = int(data.get('initial_grand_total', 0))
        # 如果是查詢2023年
        if input_year == '2023':
            initial_grand_total = 0


        # 各月份攤提金額
        installment_amount = int(data['installment_amount']) if data['installment_amount'] is not None else 0
        # 各月份[殘值]攤提金額
        resi_installment_amount = int(data['resi_installment_amount']) if data['resi_installment_amount'] is not None else 0


        # 初始化字典
        if asset_no not in data_d:
            data_d[asset_no] = {
                'account_no': data['account_no'],  # 會計編號
                'account_name': data['account_name'],  # 會計科目
                'asset_vendor': data['asset_vendor'],  # 廠商
                'asset_name': data['asset_name'],  # 設備名稱
                'depart_no': data['depart_no'],  # 部門代號
                'credit_account_no': data['credit_account_no'] if data.get('credit_account_no') is not None else '',  # 貸方編號
                'asset_qty': data['asset_qty'],  # 數量
                'asset_unit': data['asset_unit'],  # 單位
                'asset_date': data['asset_date'] if data.get('asset_date') is not None else '',  # 取得日期
                'asset_amount': int(data['asset_amount']) if data['asset_amount'] is not None else 0,  # 取得原價
                'asset_residual': int(data['asset_residual']) if data['asset_residual'] is not None else 0,  # 殘值
                'number_installment': data['number_installment'] if data['number_installment'] is not None else 0,  # 攤提期數(Y/M)
                'resi_balance_amount': resi_balance_amount,  # 殘值攤提餘額
                'initial_grand_total': initial_grand_total,  # 期初累折金額
                'monthly_installment_amounts': {}  # 每月份攤提金額
            }

        # 將當月金額加到對應的月份
        if data.get('installment_date') is not None:
            installment_month = data['installment_date'].month
        else:
            installment_month = 0
        data_d[asset_no]['monthly_installment_amounts'].setdefault(installment_month, 0)
        data_d[asset_no]['monthly_installment_amounts'][installment_month] += (installment_amount + resi_installment_amount)



    # 寫入excel
    wb = Workbook()
    wb.remove(wb.worksheets[0])
    ws = wb.create_sheet('財產目錄表')
    ws.cell(1, 1, '豆酥朋食品股份有限公司')
    # 民國
    roc_year_1 = int(input_year) - 1911
    ws.cell(2, 1, f'{roc_year_1}年財產目錄')
    ws.merge_cells('A1:AA1')
    ws.merge_cells('A2:AA2')
    alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].alignment = alignment
    ws['A2'].alignment = alignment

    topic = ['會計編號', '貸方編號', '會計科目', '廠商', '設備名稱', '部門', '數量', 
            '單位', '取得日期', '取得原價', '殘值', 'Y/M', '期初累折',
            'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 
            '本期提列數', '期末累折', '未折減餘額']
    
    # 欄位寬度
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20 
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['Y'].width = 20
    ws.column_dimensions['Z'].width = 20
    ws.column_dimensions['AA'].width = 20
    ws.column_dimensions['AB'].width = 20
    for i in range(0, len(topic)):
        ws.cell(3, i+1, topic[i]).alignment = alignment

    # 固定標題行
    ws.freeze_panes = ws['A4']

    # 將資料轉換為列表以便排序
    sorted_data = sorted(data_d.items(), key=lambda item: (item[1]['account_no'], item[1]['depart_no']))
    # 黃色樣式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    row_count = 4

    
    total_asset_amount = 0
    total_asset_residual = 0
    total_initial_grand_total = 0
    total_final_total = 0
    total_remaining_balance = 0

    # 每月份小計(條件: 會計編號 + 會計科目 + 部門)
    total_Jan = 0
    total_Feb = 0
    total_Mar = 0
    total_Apr = 0
    total_May = 0
    total_Jun = 0
    total_Jul = 0
    total_Aug = 0
    total_Sep = 0
    total_Oct = 0
    total_Nov = 0
    total_Dec = 0

    # 每月份小計(條件: 會計編號)
    monthly_totals_2 = [0] * 12

    # 依照[會計科目]跟[部門]儲存小計(底下統計需要的資料)
    monthly_totals = {}
    # 依照[貸方科目]儲存小計(最底下固定貸方科目的資料)
    credit_account_no_monthly_totals = {}

    for index, (k, v) in enumerate(sorted_data):
        # 每月份小計(條件: 會計編號)
        # 檢查是否需要插入小計行(至少要有1行account_no資料，且當account_no跟上一行account_no不同時，就插入小計，統計上一個account_no的資料)
        if index > 0 and (v['account_no'] != sorted_data[index - 1][1]['account_no']):
            # 在此處插入小計行
            row_count += 1
            ws.cell(row_count, 4, '小計')
            ws.cell(row_count, 9, total_asset_amount)        # 總資產金額
            ws.cell(row_count, 10, total_asset_residual)     # 殘值
            ws.cell(row_count, 12, total_initial_grand_total) # 期初累折金額
            for month in range(1, 13):
                ws.cell(row_count, month + 12, monthly_totals_2[month - 1]) # 1~12月小計
            ws.cell(row_count, 26, total_final_total)        # 期末累折
            ws.cell(row_count, 27, total_remaining_balance)   # 未折減餘額

            # 將整行設置為黃色
            for col in range(1, 28):
                ws.cell(row_count, col).fill = yellow_fill

            # 重置總計變數
            total_asset_amount = 0
            total_asset_residual = 0
            total_initial_grand_total = 0
            total_final_total = 0
            total_remaining_balance = 0

            # 重置每月小計
            monthly_totals_2 = [0] * 12

            row_count += 2


        # 記錄小計(條件: 會計編號 + 部門)，最底下統計用
        if index > 0 and (v['account_no'] != sorted_data[index - 1][1]['account_no'] or 
                        v['depart_no'] != sorted_data[index - 1][1]['depart_no']):
            
            monthly_totals[(sorted_data[index - 1][1]['account_no'],
                            sorted_data[index - 1][1]['account_name'],
                         sorted_data[index - 1][1]['depart_no'])] = {
                1: total_Jan,
                2: total_Feb,
                3: total_Mar,
                4: total_Apr,
                5: total_May,
                6: total_Jun,
                7: total_Jul,
                8: total_Aug,
                9: total_Sep,
                10: total_Oct,
                11: total_Nov,
                12: total_Dec,
            }
            if v['credit_account_no']:
                key = (sorted_data[index - 1][1]['credit_account_no'], sorted_data[index - 1][1]['account_name'])            
                # 如果該 credit_account_no 和 account_name 組合已經存在，則加總數據
                if key in credit_account_no_monthly_totals:
                    totals = credit_account_no_monthly_totals[key]
                    totals[1] += total_Jan
                    totals[2] += total_Feb
                    totals[3] += total_Mar
                    totals[4] += total_Apr
                    totals[5] += total_May
                    totals[6] += total_Jun
                    totals[7] += total_Jul
                    totals[8] += total_Aug
                    totals[9] += total_Sep
                    totals[10] += total_Oct
                    totals[11] += total_Nov
                    totals[12] += total_Dec
                else:
                    # 如果尚未存在，則直接創建新的條目
                    credit_account_no_monthly_totals[key] = {
                        1: total_Jan,
                        2: total_Feb,
                        3: total_Mar,
                        4: total_Apr,
                        5: total_May,
                        6: total_Jun,
                        7: total_Jul,
                        8: total_Aug,
                        9: total_Sep,
                        10: total_Oct,
                        11: total_Nov,
                        12: total_Dec,
                    }
            
            # 重置每月小計
            total_Jan = 0
            total_Feb = 0
            total_Mar = 0
            total_Apr = 0
            total_May = 0
            total_Jun = 0
            total_Jul = 0
            total_Aug = 0
            total_Sep = 0
            total_Oct = 0
            total_Nov = 0
            total_Dec = 0


        # 寫入資料
        ws.cell(row_count, 1, v['account_no']).alignment = alignment          # 會計編號
        ws.cell(row_count, 2, v['credit_account_no']).alignment = alignment   # 貸方編號
        ws.cell(row_count, 3, v['account_name']).alignment = alignment        # 會計科目
        ws.cell(row_count, 4, v['asset_vendor']).alignment = alignment        # 廠商
        ws.cell(row_count, 5, v['asset_name'])                                # 設備名稱
        ws.cell(row_count, 6, v['depart_no']).alignment = alignment           # 部門代號
        ws.cell(row_count, 7, v['asset_qty']).alignment = alignment           # 數量
        ws.cell(row_count, 8, v['asset_unit']).alignment = alignment          # 單位

        # 西元轉民國
        asset_date = v['asset_date']
        roc_year = asset_date.year - 1911
        roc_date = f"{roc_year}/{asset_date.month}/{asset_date.day}"
        ws.cell(row_count, 9, roc_date).alignment = alignment                 # 取得日期
        ws.cell(row_count, 10, v['asset_amount'])                             # 取得原價
        ws.cell(row_count, 11, v['asset_residual'])                           # 殘值
        ws.cell(row_count, 12, v['number_installment']).alignment = alignment # 攤提期數
        ws.cell(row_count, 13, v['initial_grand_total'])                      # 期初累折金額

        # 每月攤提金額
        for month in range(1, 13):
            amount = v['monthly_installment_amounts'].get(month, 0)         # 如果沒資料填入0
            ws.cell(row_count, 13 + month, '' if amount == 0 else amount)   # 把0都轉成空字串
            monthly_totals_2[month - 1] += amount
            if month == 1:
                total_Jan += amount
            elif month == 2:
                total_Feb += amount
            elif month == 3:
                total_Mar += amount
            elif month == 4:
                total_Apr += amount
            elif month == 5:
                total_May += amount
            elif month == 6:
                total_Jun += amount
            elif month == 7:
                total_Jul += amount
            elif month == 8:
                total_Aug += amount
            elif month == 9:
                total_Sep += amount
            elif month == 10:
                total_Oct += amount
            elif month == 11:
                total_Nov += amount
            elif month == 12:
                total_Dec += amount

        # 本期提列數 : 1到12月的金額總和
        total_installments = sum(v['monthly_installment_amounts'].get(month, 0) for month in range(1, 13))
        ws.cell(row_count, 26, total_installments)

        # 期末累折
        final_total = v['initial_grand_total'] + total_installments
        ws.cell(row_count, 27, final_total)

        # 未折減餘額
        remaining_balance = v['asset_amount'] - final_total
        remaining_balance = max(0, remaining_balance)
        ws.cell(row_count, 28, remaining_balance)

        # 累計各項金額
        total_asset_amount += v['asset_amount']
        total_asset_residual += v['asset_residual']
        total_initial_grand_total += v['initial_grand_total']
        total_final_total += final_total
        total_remaining_balance += remaining_balance
        row_count += 1


    # 在最後一個會計編號後插入小計行
    if total_asset_amount > 0:  # 確保有需要的小計
        row_count += 1
        ws.cell(row_count, 4, '小計')
        ws.cell(row_count, 9, total_asset_amount)        # 總資產金額
        ws.cell(row_count, 10, total_asset_residual)     # 殘值
        ws.cell(row_count, 12, total_initial_grand_total) # 期初累折金額
        for month in range(1, 13):
            ws.cell(row_count, month + 12, monthly_totals_2[month - 1]) # 1~12月小計
        ws.cell(row_count, 26, total_final_total)        # 期末累折
        ws.cell(row_count, 27, total_remaining_balance)   # 未折減餘額
        # 將整行設置為黃色
        for col in range(1, 28):
            ws.cell(row_count, col).fill = yellow_fill

        monthly_totals[(v['account_no'], v['account_name'], v['depart_no'])] = {
            1: total_Jan,
            2: total_Feb,
            3: total_Mar,
            4: total_Apr,
            5: total_May,
            6: total_Jun,
            7: total_Jul,
            8: total_Aug,
            9: total_Sep,
            10: total_Oct,
            11: total_Nov,
            12: total_Dec,
        }

        if v['credit_account_no']:
            key = (sorted_data[index - 1][1]['credit_account_no'], sorted_data[index - 1][1]['account_name'])
            if key in credit_account_no_monthly_totals:
                totals = credit_account_no_monthly_totals[key]
                totals[1] += total_Jan
                totals[2] += total_Feb
                totals[3] += total_Mar
                totals[4] += total_Apr
                totals[5] += total_May
                totals[6] += total_Jun
                totals[7] += total_Jul
                totals[8] += total_Aug
                totals[9] += total_Sep
                totals[10] += total_Oct
                totals[11] += total_Nov
                totals[12] += total_Dec
            else:
                credit_account_no_monthly_totals[key] = {
                    1: total_Jan,
                    2: total_Feb,
                    3: total_Mar,
                    4: total_Apr,
                    5: total_May,
                    6: total_Jun,
                    7: total_Jul,
                    8: total_Aug,
                    9: total_Sep,
                    10: total_Oct,
                    11: total_Nov,
                    12: total_Dec,
                }

    row_count += 2
    topic = [
            '1月', '2月', '3月', '4月', '5月', '6月', 
            '7月', '8月', '9月', '10月', '11月', '12月', ]
    for i in range(0, len(topic)):
        ws.cell(row_count, i+13, topic[i]).alignment = alignment
    row_count += 1
    ws.cell(row_count, 12, '填傳編→').alignment = alignment
    ws.cell(row_count, 25, '合計').alignment = alignment
    row_count += 1
    monthly_totals_sum = [0] * 13
    color_dict = {}
    # 寫入每個會計科目的每月小計
    for (account_no, account_name, depart_no), totals in monthly_totals.items():
        # 檢查是否已為該 account_no 分配顏色
        if account_no not in color_dict:
            random_color = get_random_color()
            color_dict[account_no] = random_color

        # 設置顏色
        font_color = Font(color=color_dict[account_no])

        ws.cell(row_count, 10, account_no).alignment = alignment
        ws.cell(row_count, 10, account_no).font = font_color
        ws.cell(row_count, 11, f'{roc_year_1}年{account_name}提列折舊').font = font_color
        ws.cell(row_count, 12, depart_no).alignment = alignment 
        ws.cell(row_count, 12, depart_no).font = font_color

        for month in range(1, 13):
            amount = totals[month]
            ws.cell(row_count, 12 + month, amount).font = font_color
            monthly_totals_sum[month - 1] += amount


        total_month = sum(totals[month] for month in range(1, 13))
        ws.cell(row_count, 25, total_month)
        ws.cell(row_count, 25, total_month).font = font_color
        monthly_totals_sum[12] += total_month
        row_count += 1

    # 總計
    for col in range(12, 26):
        ws.cell(row_count, col).fill = yellow_fill
    ws.cell(row_count, 12, '總計').alignment = alignment
    for month in range(0, 12):
        ws.cell(row_count, 13 + month, monthly_totals_sum[month])
    ws.cell(row_count, 25, monthly_totals_sum[12])

    row_count += 3

    sorted_totals = sorted(credit_account_no_monthly_totals.items(), key=lambda x:x[0][0])
    #   sorted_totals = [
    #     (('A001', 'Account 1'), {1: 100, 2: 200, 3: 300}),
    #     (('A003', 'Account 3'), {1: 120, 2: 220, 3: 320}),
    #     (('B002', 'Account 2'), {1: 150, 2: 250, 3: 350})
    # ]
    credit_account_no_monthly_totals_sum = [0] * 13
    for (credit_account_no, account_name), totals in sorted_totals:
        ws.cell(row_count, 10, credit_account_no).alignment = alignment
        ws.cell(row_count, 11, f'{roc_year_1}年{account_name}提列折舊')

        for month in range(1, 13):
            amount = totals[month]
            ws.cell(row_count, 12 + month, amount)
            credit_account_no_monthly_totals_sum[month - 1] += amount


        total_month = sum(totals[month] for month in range(1, 13))
        ws.cell(row_count, 25, total_month)
        credit_account_no_monthly_totals_sum[12] += total_month
        row_count += 1

    # 總計
    for col in range(12, 26):
        ws.cell(row_count, col).fill = yellow_fill
    ws.cell(row_count, 12, '總計').alignment = alignment
    for month in range(0, 12):
        ws.cell(row_count, 13 + month, credit_account_no_monthly_totals_sum[month])
    ws.cell(row_count, 25, credit_account_no_monthly_totals_sum[12])


    # 設定數字格式
    for col in [i for i in range(9, 28) if i not in [11]]:      # 需要格式化的列
        for row in range(4, row_count + 1):  # 從第4行到當前行
            ws.cell(row, col).number_format = '#,##0'  # 千分位格式
    
    # 如果有指定月份
    if input_month:
        # 先隱藏全部月份欄位
        for month in range(1, 13):  # 隱藏1到12
            ws.column_dimensions[get_column_letter(12 + month)].hidden = True

        # 解除隱藏input_month對應的欄位
        ws.column_dimensions[get_column_letter(12 + int(input_month))].hidden = False  


    with NamedTemporaryFile(delete=False) as tmp:
        wb.save(tmp.name)
        output = BytesIO(tmp.read())
    export_file = HttpResponse(output, content_type='application/vnd.ms-excel')
    # export_file['Content-Disposition'] = f'attachment; filename={roc_year_1}年財產目錄.xlsx'
    return export_file

def get_random_color():
    # 生成隨機顏色的 RGB
    r = random.randint(0, 180)
    g = random.randint(0, 180)
    b = random.randint(0, 180)
    return f"{r:02X}{g:02X}{b:02X}"



"""
App 進入點
"""


class Acc3003(TemplateView):
    template_name = 'acc3003.html'

    def get(self, request, *args, **kwargs):
        if request_app_auth(request, 'accounting', 'acc3003') is None:
            return HttpResponseRedirect('/')
        username = request.session['username']

        template_data = dict(
            facing_d=dict(username=username),
        )

        return render(request, self.template_name, template_data)

    def post(self, request):
        username = request.session['username']

        data_form = get_form_post_data(request)
        op = data_form['fmTextOP']
        data_ld = list()
        outline_d = dict()
        if op == 'GO':
            data_ld = acc3003_main(data_form)
            return data_ld

        template_data = dict(
            data_form=data_form,
            outline_d=outline_d,
            facing_d=dict(username=username),
        )

        return render(request, self.template_name, template_data)
